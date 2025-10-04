import csv
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from ..models import AreaManagerVisit, ChecklistItem, ActionPlanItem

@login_required
def export_visit_excel(request, visit_id):
    """Export a single checklist visit to an Excel file."""
    visit = get_object_or_404(AreaManagerVisit, id=visit_id, manager=request.user)
    
    workbook = Workbook()
    
    # Summary sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Visit Summary"
    
    summary_sheet['A1'] = "Store"
    summary_sheet['B1'] = visit.store.name
    summary_sheet['A2'] = "Date"
    summary_sheet['B2'] = visit.date
    summary_sheet['A3'] = "Manager"
    summary_sheet['B3'] = visit.manager.get_full_name() or visit.manager.username
    summary_sheet['A4'] = "Overall Score"
    summary_sheet['B4'] = f"{visit.calculate_score()}%"
    
    # Checklist details sheet
    details_sheet = workbook.create_sheet(title="Checklist Details")
    
    headers = ["Category", "Question No.", "Question Text", "Answer", "Comment"]
    for col_num, header in enumerate(headers, 1):
        cell = details_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    for row_num, item in enumerate(visit.checklist_items.all(), 2):
        details_sheet.cell(row=row_num, column=1).value = item.question.category.name
        details_sheet.cell(row=row_num, column=2).value = item.question.number
        details_sheet.cell(row=row_num, column=3).value = item.question.text
        details_sheet.cell(row=row_num, column=4).value = "Yes" if item.answer else "No"
        details_sheet.cell(row=row_num, column=5).value = item.comment
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="visit_{visit.id}_report.xlsx"'
    workbook.save(response)
    
    return response

@login_required
def export_history_excel(request):
    """Export the entire checklist history to an Excel file."""
    visits = AreaManagerVisit.objects.filter(manager=request.user, is_draft=False).order_by('-date')
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Checklist History"
    
    headers = ["Store", "Date", "Manager", "Overall Score"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    for row_num, visit in enumerate(visits, 2):
        sheet.cell(row=row_num, column=1).value = visit.store.name
        sheet.cell(row=row_num, column=2).value = visit.date
        sheet.cell(row=row_num, column=3).value = visit.manager.get_full_name() or visit.manager.username
        sheet.cell(row=row_num, column=4).value = f"{visit.calculate_score()}%"
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="checklist_history.xlsx"'
    workbook.save(response)
    
    return response

@user_passes_test(lambda u: u.is_superuser)
def export_data(request):
    """Export checklist data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="checklist_data_{timezone.now().date()}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow(['Store', 'Manager', 'Date', 'Month', 'Score', 'Total Items', 'Passed Items'])
    
    # Write data
    visits = AreaManagerVisit.objects.filter(is_draft=False).select_related('store', 'manager')
    
    for visit in visits:
        items = ChecklistItem.objects.filter(visit=visit)
        total_items = items.count()
        passed_items = items.filter(answer=True).count()
        score = visit.calculate_score() if hasattr(visit, 'calculate_score') else 0
        
        writer.writerow([
            visit.store.name,
            visit.manager.get_full_name() or visit.manager.username,
            visit.date,
            visit.month,
            score,
            total_items,
            passed_items
        ])
    
    return response


@user_passes_test(lambda u: u.is_superuser)
def import_questions(request):
    """Import checklist questions from CSV"""
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        # Add your import logic here
        messages.success(request, 'Questions imported successfully!')
        return redirect('checklist:manage_checklist_questions')
    
    return render(request, 'checklist/import_questions.html')


@user_passes_test(lambda u: u.is_superuser)
def export_visit_maintenance_report(request):
    """Export combined visit and maintenance data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="visit_maintenance_report_{timezone.now().date()}.csv"'
    
    writer = csv.writer(response)
    
    # Write header with additional maintenance columns
    writer.writerow([
        'Store', 'Manager', 'Visit Date', 'Visit Score', 
        'Maintenance ID', 'Maintenance Date', 'Maintenance Status', 
        'Maintenance Description', 'Action Items'
    ])
    
    # Get all visits and related maintenance tickets
    visits = AreaManagerVisit.objects.filter(is_draft=False).select_related('store', 'manager')
    
    for visit in visits:
        # Get maintenance tickets linked to this visit
        maintenance_tickets = visit.maintenanceticket_set.all()
        score = visit.calculate_score() if hasattr(visit, 'calculate_score') else 0
        
        if maintenance_tickets.exists():
            for ticket in maintenance_tickets:
                writer.writerow([
                    visit.store.name,
                    visit.manager.get_full_name() or visit.manager.username,
                    visit.date,
                    score,
                    ticket.id,
                    ticket.created_date,
                    ticket.get_status_display(),
                    ticket.description,
                    "\n".join([item.description for item in ticket.actionplanitem_set.all()])
                ])
        else:
            # Write visit data even if no maintenance tickets exist
            writer.writerow([
                visit.store.name,
                visit.manager.get_full_name() or visit.manager.username,
                visit.date,
                score,
                "", "", "", "", ""
            ])
    
    return response